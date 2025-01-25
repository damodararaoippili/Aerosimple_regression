import openpyxl
file = "C:\\Users\\damod\\PycharmProjects\\Aerosimple\\data\\Aerosimpletestdata.xlsx"
workbook = openpyxl.load_workbook(file)
def read_logindata():
    # Read loginpage data from the sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook['Login details']
        rows = sheet.max_row
        testdata = {}
        for row in range(2, rows + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            testdata[key] = value
        return testdata
