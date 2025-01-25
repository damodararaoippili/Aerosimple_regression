import openpyxl
File = "C:\\Users\\damod\\PycharmProjects\\Aerosimple\\data\\Aerosimpletestdata.xlsx"
workbook = openpyxl.load_workbook(File)
def get_testcase_data(sheet_name, testcase_id):
    if sheet_name in workbook.sheetnames:
        current_sheet = workbook[sheet_name]
        rows = current_sheet.max_row
        columns = current_sheet.max_column

        for row in range(1, rows + 1):
            testcase = current_sheet.cell(row=row, column=1).value
            if testcase == testcase_id:
                header_row = row - 1

                headers = []
                for col in range(1, columns + 1):
                    header_value = current_sheet.cell(row=header_row, column=col).value
                    if header_value:
                        headers.append(header_value)

                test_case_data = {}
                for col in range(1, columns + 1):
                    if col <= len(headers):
                        header = headers[col - 1]
                        value = current_sheet.cell(row=row, column=col).value
                        if value:
                            test_case_data[header] = value

                return test_case_data

        return None
    else:
        print(f"Sheet '{sheet_name}' does not exist.")
        return None

test_case_data = get_testcase_data('AWO', 'TC_AW005')
print(test_case_data)
